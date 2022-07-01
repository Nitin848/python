import openpyxl
import openpyxl as openpyxl

wb = openpyxl.load_workbook("C:\\Users\\nitin\\OneDrive\\Desktop\\information.xlsx")
print(type(wb))

'''sheets=wb.sheetnames
print(sheets)

print(wb.active.title)

sh1=wb['names']
print(type(sh1))

c=sh1.cell(row=17,column=2)
print(c)
print(c.value)
data=sh1['B6'].value
(data)

(sh1.cell(row=18,column=2).value)'''
print(wb.get_sheet_by_name('names').cell(row=18,column=2).value)